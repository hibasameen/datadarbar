#!/usr/bin/env python3
"""Parse PBS industry statistics (QIM / LSM / CMI) into warehouse + app JSON.

Inputs (../../industry_statistics/):
  qim/Trend-sheet.xlsx           monthly overall QIM Jul-2016 -> latest
  qim/Table-2-may-2026.xlsx      monthly sector indices FY2025-26
  qim/Table-3-2024-25.xlsx       monthly sector indices FY2024-25
  table2_new/Table-2-YYYY-YY.pdf annual+monthly sector indices, base 2015-16 (2015-16..2023-24)
  table2_old/Table-2-YYYY-YY.pdf same, base 2005-06 (2005-06..2021-22)
  reports/CMI_2015-16_report.pdf summary tables 3.2/3.3/3.5/3.13

Outputs:
  ../lsm_qim.parquet, ../lsm_sector_indices.parquet (warehouse)
  econ_data/industry.json (app bundle input)

Notes: PDF numbers may contain stray spaces ("1, 093, 235", "84. 9").
Sector rows are matched against canonical sector names taken from the xlsx.
"""
import json, re, subprocess, sys, glob, os
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[2] / 'industry_statistics'
OUT_WH = Path(__file__).resolve().parents[1]
OUT_JSON = Path(__file__).resolve().parents[1] / 'econ_data'
OUT_JSON.mkdir(exist_ok=True)

MONTH_ABBR = ['Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar','Apr','May','Jun']

# ---------------- helpers ----------------
def clean_num(tok):
    tok = tok.replace(',', '').replace(' ', '')
    try:
        return float(tok)
    except ValueError:
        return None

def fix_number_spaces(line):
    # "1, 093, 235" -> "1,093,235" ; "84. 9" -> "84.9"
    prev = None
    while prev != line:
        prev = line
        line = re.sub(r'(\d)[,.]\s+(\d)', lambda m: m.group(0).replace(' ', ''), line)
    return line

def nums_in(line):
    return [clean_num(t) for t in re.findall(r'-?[\d][\d,]*\.?\d*', line)]

def canon(name):
    n = re.sub(r'[^a-z]+', ' ', name.lower()).strip()
    n = re.sub(r'\b(manufacturing|manufacture|of|the)\b', '', n).strip()
    return re.sub(r'\s+', ' ', n)

# ---------------- 1. QIM trend sheet ----------------
def parse_trend():
    wb = openpyxl.load_workbook(ROOT / 'qim/Trend-sheet.xlsx', data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(min_row=6, values_only=True):
        if r[0] is None or r[1] is None:
            continue
        d = r[0]
        if not hasattr(d, 'year'):
            continue
        # fiscal month label
        rows.append(dict(month=f'{d.year:04d}-{d.month:02d}', qim=round(float(r[1]), 2),
                         mom=round(float(r[2]), 2) if r[2] is not None else None,
                         yoy=round(float(r[3]), 2) if r[3] is not None else None,
                         cum_qim=round(float(r[4]), 2) if r[4] is not None else None,
                         cum_chg=round(float(r[5]), 2) if r[5] is not None else None))
    return rows

# ---------------- 2. sector list + monthly indices from xlsx ----------------
def parse_sector_xlsx(path, fy_start_year):
    """Sector rows: col B empty, col A = sector name, col D = weight, cols E.. = monthly indices."""
    wb = openpyxl.load_workbook(ROOT / path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sectors = {}
    for r in ws.iter_rows(min_row=6, values_only=True):
        a, b, c, d = r[0], r[1], r[2], r[3]
        if a is None or d is None:
            continue
        if b is None and isinstance(a, str) and not re.match(r'^[\d\s\-]+$', a):
            name = a.strip()
            monthly = {}
            for i, mv in enumerate(r[4:16]):
                if mv is None or (isinstance(mv, str) and not mv.strip()):
                    continue
                y = fy_start_year if i < 6 else fy_start_year + 1
                try:
                    fv = float(mv)
                    if fv == 0:      # 0 in a sector index row = month not yet reported
                        continue
                    monthly[f'{y:04d}-{[7,8,9,10,11,12,1,2,3,4,5,6][i]:02d}'] = round(fv, 2)
                except (TypeError, ValueError):
                    pass
            annual = None
            v = r[16]
            try:
                annual = round(float(v), 2) if v is not None else None
            except (TypeError, ValueError):
                annual = None
            sectors[name] = dict(weight=round(float(d), 3), monthly=monthly, annual=annual)
    return sectors

# ---------------- 3. Table-2 PDFs ----------------
UNIT_WORDS = {'mt', 'l', 'litres', 'sq m', 'nos', 'mil nos', 'pairs', 'sq ft', 'unit', 'quantity'}

def _name_and_nums(line):
    """Leading alpha fragment (stops at first digit) + the numbers that follow it.

    An S.No RANGE prefix like "24-35 Coke & Petroleum Products..." is allowed
    (old-base sector rows carry one); its numbers are excluded from vals."""
    m = re.match(r"^\s*(?:\d+\s*-\s*\d+\s+)?([A-Za-z][^\d]*?)\s*(?=\d|$)", line)
    if not m:
        return '', [v for v in nums_in(line) if v is not None]
    name = m.group(1).strip()
    # drop unit tokens glued at the end of the fragment
    name = re.sub(r"\s*(MT|000 L(itres)?|Mil\. Nos\.|`?000'? ?(Litres|L|Nos\.?|Sq ?M|Sq ?Ft|Dozen|Pairs))\s*$", '', name).strip()
    vals = [v for v in nums_in(line[m.end():]) if v is not None]
    return name, vals

def _match_canon(cn, canon_map):
    if cn in canon_map:
        return canon_map[cn]
    if len(cn) >= 6:
        cands = [k for k in canon_map if k.startswith(cn) or cn.startswith(k)]
        if len(cands) == 1:
            return canon_map[cands[0]]
        if cands:  # prefer longest common
            return canon_map[max(cands, key=len)]
    return None

def parse_table2_pdf(path, canonical, fy_label, fallback_weights=None):
    """Return {canonical_sector: {weight, annual, monthly{}}} for one fiscal year PDF.

    Handles: names wrapped onto the values line, weights glued to names,
    deep indentation, and letter-spaced text in some old-base files.
    """
    txt = subprocess.run(['pdftotext', '-layout', str(path), '-'],
                         capture_output=True, text=True).stdout
    lines = [fix_number_spaces(l) for l in txt.split('\n')]
    fy0 = int(fy_label[:4])
    out = {}
    canon_map = {canon(k): k for k in canonical}
    has_weight = bool(re.search(r'We ?i ?g ?h ?t', txt[:3000]))

    def commit(name, vals):
        cn = canon(name)
        key = _match_canon(cn, canon_map)
        if not key or key in out:
            return False
        if not has_weight:
            if len(vals) < 13:
                return False
            monthly_vals, annual = vals[0:12], vals[12]
            weight = (fallback_weights or {}).get(key)
        elif len(vals) == 13:
            weight, monthly_vals, annual = vals[0], vals[1:13], None
        elif len(vals) >= 14:
            weight, monthly_vals, annual = vals[0], vals[1:13], vals[13]
        else:
            return False
        if weight is not None and weight > 100:   # production quantities, not an index row
            return False
        if any(v < 0 or v > 5000 for v in monthly_vals):
            return False
        monthly = {}
        for k, mv in enumerate(monthly_vals):
            y = fy0 if k < 6 else fy0 + 1
            monthly[f'{y:04d}-{[7,8,9,10,11,12,1,2,3,4,5,6][k]:02d}'] = mv
        out[key] = dict(weight=weight, monthly=monthly, annual=annual)
        return True

    i = 0
    while i < len(lines):
        name, vals = _name_and_nums(lines[i])
        if name and canon(name):
            if len(vals) >= 13:
                commit(name, vals)
            elif len(vals) <= 1 and i + 1 < len(lines):
                # name possibly wrapped; values (and a name fragment) on next line
                name2, vals2 = _name_and_nums(lines[i + 1])
                merged = vals + vals2 if len(vals) == 1 else vals2
                committed = False
                if len(merged) >= 13:
                    frag = '' if canon(name2) in ('', 'mt') or set(name2.lower().split()) <= UNIT_WORDS else ' ' + name2
                    if commit(name + frag, merged):
                        i += 1
                        committed = True
                if not committed and i > 0:
                    # ...or on the line ABOVE the name (seen in 2022-23 Paper & Board)
                    name0, vals0 = _name_and_nums(lines[i - 1])
                    if not name0 and len(vals0) >= 13:
                        commit(name, vals0)
        i += 1
    return out

# ---------------- 4. CMI summary tables ----------------
CMI_DIVS = {  # PSIC division -> short label
 '10':'Food','11':'Beverages','12':'Tobacco','13':'Textiles','14':'Wearing apparel',
 '15':'Leather','16':'Wood','17':'Paper','18':'Printing','19':'Coke & petroleum',
 '20':'Chemicals','21':'Pharmaceuticals','22':'Rubber & plastics','23':'Non-metallic minerals',
 '24':'Basic metals','25':'Fabricated metal','26':'Electronics','27':'Electrical equipment',
 '28':'Machinery','29':'Motor vehicles','30':'Other transport','31':'Furniture','32':'Other mfg',
 'OD':'All other'}

def parse_cmi():
    txt = subprocess.run(['pdftotext', '-layout', str(ROOT / 'reports/CMI_2015-16_report.pdf'), '-'],
                         capture_output=True, text=True).stdout
    lines = txt.split('\n')
    def section(start_pat, end_pat, lo=0, hi=3000):
        s = e = None
        for i, l in enumerate(lines[lo:hi], lo):
            if s is None and re.search(start_pat, l):
                s = i
            elif s is not None and re.search(end_pat, l):
                e = i
                break
        return lines[s:e] if s is not None else []

    # Table 3.2 aggregates across censuses
    agg = {}
    sec = section(r'Table 3\.2:', r'Number of Reporting Establishments')
    joined = ' '.join(sec)
    for label, key in [('Output at producer', 'output'), ('Industrial Cost', 'industrial_cost')]:
        pass  # aggregate rows wrap; simpler to regex the whole section per row id
    # rows of interest with 5 census values
    def grab(pat):
        m = re.search(pat + r'[\s\S]{0,220}?([\d,]{4,})\s+([\d,]{4,})\s+([\d,]{4,})\s+([\d,]{4,})\s+([\d,]{4,})', joined)
        return [clean_num(g) for g in m.groups()] if m else None
    agg['censuses'] = ['1990-91', '1995-96', '2000-01', '2005-06', '2015-16']
    agg['output_producer_prices'] = grab(r'Output at producer')
    agg['cva_market_prices'] = grab(r'Census Value Added\s+5?\s*\(Market Prices\)')
    if agg['cva_market_prices'] is None:
        agg['cva_market_prices'] = grab(r'\(Market Prices\)')
    agg['gdp_contribution_bp'] = grab(r'Contribution to GDP\s+16?\s*\(Basic Prices\)')
    if agg['gdp_contribution_bp'] is None:
        agg['gdp_contribution_bp'] = grab(r'\(Basic Prices\) \(14-')

    def div_table(start_pat, end_pat, ncols):
        sec = section(start_pat, end_pat)
        rows = {}
        for l in sec:
            m = re.match(r'\s*(1[0-9]|2[0-9]|3[0-2]|OD)\s+(.+)', l)
            if not m:
                continue
            div = m.group(1)
            vals = [v for v in nums_in(fix_number_spaces(m.group(2))) if v is not None]
            if len(vals) >= ncols:
                rows[div] = vals[:ncols]
        return rows

    # establishments: 2005-06, 2015-16, %05, %15, change
    est = div_table(r'Table 3\.3:', r'Employment', 4)
    # employment: 2005-06, 2015-16, %05, %15, change, acgr
    emp = div_table(r'Table 3\.5:', r'Value of Production', 4)
    # GVA weights: 2005-06, 2015-16 non-adj, adjusted
    wts = div_table(r'Table 3\.13:', r'Labour Productivity', 3)
    return dict(aggregates=agg,
                establishments={d: dict(label=CMI_DIVS.get(d, d), v2005=v[0], v2015=v[1]) for d, v in est.items()},
                employment={d: dict(label=CMI_DIVS.get(d, d), v2005=v[0], v2015=v[1]) for d, v in emp.items()},
                gva_weights={d: dict(label=CMI_DIVS.get(d, d), w2005=v[0], w2015=v[2]) for d, v in wts.items()})

# ---------------- main ----------------
def main():
    trend = parse_trend()
    fy2526 = parse_sector_xlsx('qim/Table-2-may-2026.xlsx', 2025)
    fy2425 = parse_sector_xlsx('qim/Table-3-2024-25.xlsx', 2024)
    canonical = sorted(set(fy2526) | set(fy2425))
    print(f'trend months: {len(trend)}  canonical sectors: {len(canonical)}')
    for s in canonical:
        print('  ', s, fy2526.get(s, {}).get('weight'))

    fw = {s: d['weight'] for s, d in fy2526.items()}
    new_base = {}
    for f in sorted(glob.glob(str(ROOT / 'table2_new/Table-2-*.pdf'))):
        fy = re.search(r'(\d{4}-\d{2})', os.path.basename(f)).group(1)
        new_base[fy] = parse_table2_pdf(f, canonical, fy, fallback_weights=fw)
        print(f'new {fy}: {len(new_base[fy])} sectors')
    new_base['2024-25'] = fy2425
    new_base['2025-26'] = fy2526

    # QIM total row absent from 2021-22/2022-23 PDFs: fill from trend sheet
    # (annual = cumulative QIM at June of the fiscal year)
    trend_by_month = {t['month']: t for t in trend}
    for fy in new_base:
        if 'QIM' not in new_base[fy]:
            jun = f'{int(fy[:4])+1:04d}-06'
            if jun in trend_by_month:
                mo = {m['month']: m['qim'] for m in trend if fy[:4] <= m['month'][:4] <= str(int(fy[:4])+1)}
                monthly = {k: v for k, v in mo.items()
                           if (k >= f'{fy[:4]}-07' and k <= f'{int(fy[:4])+1}-06')}
                new_base[fy]['QIM'] = dict(weight=78.37, monthly=monthly,
                                           annual=trend_by_month[jun]['cum_qim'])

    # old base: discover names on first file, then parse all
    oldnames = ['Textile', 'Food, Beverages & Tobacco', 'Coke & Petroleum Products',
                'Pharmaceuticals', 'Chemicals', 'Automobiles', 'Fertilizers',
                'Electronics', 'Leather Products', 'Paper & Board', 'Engineering Products',
                'Rubber Products', 'Non-Metallic Mineral Products', 'Wood Products',
                'Iron & Steel products', 'Iron & Steel Products', 'QIM',
                'Non Metalic Mineral Products', 'Paper & Board']
    old_base = {}
    for f in sorted(glob.glob(str(ROOT / 'table2_old/Table-2-*.pdf'))):
        fy = re.search(r'(\d{4}-\d{2})', os.path.basename(f)).group(1)
        old_base[fy] = parse_table2_pdf(f, oldnames, fy)
        print(f'old {fy}: {len(old_base[fy])} sectors  {sorted(old_base[fy])[:3]}...')

    cmi = parse_cmi()
    print('CMI:', {k: (len(v) if isinstance(v, dict) else 'ok') for k, v in cmi.items()})

    out = dict(trend=trend, canonical=canonical,
               new_base=new_base, old_base=old_base, cmi=cmi,
               meta=dict(source='PBS Industry Statistics (pbs.gov.pk/industry-2)',
                         qim_base='2015-16=100', old_base='2005-06=100',
                         downloaded='2026-08-09'))
    (OUT_JSON / 'industry.json').write_text(json.dumps(out))
    print('wrote econ_data/industry.json', (OUT_JSON / 'industry.json').stat().st_size // 1024, 'KB')

    # warehouse parquet
    pd.DataFrame(trend).to_parquet(OUT_WH / 'lsm_qim.parquet', index=False)
    rows = []
    for base, data in [('2015-16', new_base), ('2005-06', old_base)]:
        for fy, sectors in data.items():
            for s, d in sectors.items():
                rows.append(dict(base=base, fy=fy, sector=s, weight=d['weight'],
                                 annual_index=d.get('annual')))
                for mo, v in d['monthly'].items():
                    rows.append(dict(base=base, fy=fy, sector=s, weight=d['weight'],
                                     month=mo, monthly_index=v))
    pd.DataFrame(rows).to_parquet(OUT_WH / 'lsm_sector_indices.parquet', index=False)
    print('wrote parquet:', len(rows), 'rows')

if __name__ == '__main__':
    main()
